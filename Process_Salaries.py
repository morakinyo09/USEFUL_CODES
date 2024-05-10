#Application to increment the salary of department staffs by 25%
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

workbook = xl.load_workbook("test_department.xlsx")
sheet = workbook["Sheet1"]
#cell = sheet["c3"]
#print("Cell value is ", cell.value)

percent_increment = 0.25
old_salary_col = 3
increment_col = 4
new_salary_col = 5
salary_start_row = 3
salary_end_row = sheet.max_row + 1


for row in range(salary_start_row, sheet.max_row + 1):
    old_salary_cell = sheet.cell(row, old_salary_col)
    old_salary = old_salary_cell.value
    increment = old_salary * percent_increment
    new_salary = old_salary + increment
    increment_cell = sheet.cell(row, increment_col)
    increment_cell.value = increment
    new_salary_cell = sheet.cell(row, new_salary_col)
    new_salary_cell.value = new_salary

    print("salary = ", old_salary, ", increment = ", increment, "new = ", new_salary )

#workbook.save("processed_salaries.xlsx")

chart_values = Reference(sheet, min_row = salary_start_row, max_row = salary_end_row, min_col = new_salary_col, max_col = new_salary_col )

chart = BarChart()
chart.add_data( chart_values )
sheet.add_chart(chart, "g2")

workbook.save("processed_test_department.xlsx")

